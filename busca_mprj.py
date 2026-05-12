import requests
import pandas as pd
import smtplib
import os
import time
import fitz  # PyMuPDF
from email.message import EmailMessage
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURAÇÕES ---
EMAIL_DESTINO = "renan.barros@mprj.mp.br"
EMAIL_REMETENTE = "renan.help@gmail.com" 
SENHA_APP = "saty tgmz rzrz yrai" 
GEMINI_KEY = os.getenv("GEMINI_API_KEY")

def extrair_dados_com_ia(caminho_pdf):
    tempo_processamento = 0
    status_ia = "Nao iniciado"
    
    try:
        print("Lendo PDF localmente com PyMuPDF (Estrategia Sniper)...")
        doc = fitz.open(caminho_pdf)
        texto_alvo = ""
        capturando = False
        paginas_capturadas = 0
        
        # O Robo procura a pagina exata e captura ela + 2 paginas seguintes
        for pagina in doc:
            texto_pag = pagina.get_text("text")
            
            if "CONCURSO DE REMOÇÃO" in texto_pag.upper() and "PROMOTOR" in texto_pag.upper():
                capturando = True
                
            if capturando:
                texto_alvo += texto_pag + "\n"
                paginas_capturadas += 1
                
                # Pega no maximo 3 paginas para nao confundir a IA
                if paginas_capturadas >= 3:
                    break
        doc.close()

        if not texto_alvo:
            return [], "Falha: Secao nao encontrada no PDF", 0

        print(f"Secao isolada! Comunicando DIRETAMENTE com a API do Gemini...")
        
        prompt = f"""
        Você é um analista de dados especialista em Diários Oficiais.
        Abaixo está o trecho exato onde ocorre o "CONCURSO DE REMOÇÃO PARA PROMOTOR DE JUSTIÇA".

        Sua missão é extrair as vagas listadas.
        - Identifique os itens numerados (ex: 4.1, 4.2).
        - Identifique o Órgão (Nome da Promotoria).
        - Identifique o Critério (Antiguidade ou Merecimento).
        - Identifique a Origem da vaga (ex: decorrente da promoção de Fulano).

        SAÍDA OBRIGATÓRIA (Separada por ponto e vírgula):
        Item;Órgão;Critério;Origem da Vaga

        Importante: Retorne APENAS as linhas formatadas com as vagas.
        Se não houver vagas de remoção, responda apenas: VAZIO.

        TEXTO PARA ANÁLISE:
        {texto_alvo}
        """
        
        headers = {'Content-Type': 'application/json'}
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.1}
        }
        
        # Testando direto com o modelo Flash mais robusto
        url_api = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_KEY}"
        
        inicio_ia = time.time()
        response = requests.post(url_api, headers=headers, json=payload)
        fim_ia = time.time()
        tempo_processamento = round(fim_ia - inicio_ia, 2)
        
        if response.status_code == 200:
            status_ia = "Sucesso na comunicacao"
            dados_json = response.json()
            try:
                res = dados_json['candidates'][0]['content']['parts'][0]['text'].strip()
            except KeyError:
                return [], "Erro ao interpretar resposta da API", tempo_processamento
            
            if "VAZIO" in res or ";" not in res:
                return [], status_ia, tempo_processamento
                
            linhas = [l.strip() for l in res.split('\n') if ';' in l]
            return [linha.split(';') for linha in linhas], status_ia, tempo_processamento
        else:
            erro_msg = f"Erro {response.status_code}: {response.text}"
            print(erro_msg)
            return [], f"Erro na API ({response.status_code})", tempo_processamento
        
    except Exception as e:
        print(f"Erro no processamento da IA: {e}")
        return [], f"Erro critico: {str(e)}", tempo_processamento

def formatar_excel(dados, arquivo, data_do):
    df = pd.DataFrame(dados, columns=["Item", "Órgão", "Critério", "Origem da Vaga (Decorrente de)"])
    with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=2, sheet_name='Vagas')
        ws = writer.sheets['Vagas']
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Resultados encontrados no DOeMPRJ de {data_do}"
        ws['A1'].font = Font(size=14, bold=True, color="2F5597")
        ws['A1'].alignment = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 55
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=3, max_row=len(dados)+3):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')

def enviar_email(data_do, url_pdf, localizado, status_dl, status_ia, tem_dados, qtd_vagas=0, tempo_ia=0, tamanho_kb=0, arquivo_excel=None, arquivo_pdf=None):
    msg = EmailMessage()
    msg['From'] = EMAIL_REMETENTE
    msg['To'] = EMAIL_DESTINO
    msg['Subject'] = f"Monitoramento DOeMPRJ - {data_do}"
    
    status_arquivo = "Localizado" if localizado else "Nao localizado"
    endereco_url = url_pdf if localizado else "Nao localizado"
    
    if tem_dados:
        resultado_texto = f"Sucesso. {qtd_vagas} vagas de remocao extraidas e informadas no arquivo em anexo."
    else:
        resultado_texto = "Dados de remocao nao encontrados."

    # Formatação exata do e-mail
    corpo = (
        f"Pesquisa realizada para o Diario Oficial de {data_do}.\n"
        f"{resultado_texto}\n\n"
        f"-------------------------------------------------\n"
        f"Relatorio de Execucao - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
        f"-------------------------------------------------\n\n"
        f"Arquivo DOe: {status_arquivo}\n"
        f"Endereco URL: {endereco_url}\n"
        f"Status do Download: {status_dl} (Tamanho: {tamanho_kb} KB)\n"
        f"Comunicacao com IA: {status_ia}\n"
        f"Tempo de Leitura da IA: {tempo_ia} segundos\n"
    )
    msg.set_content(corpo)

    if arquivo_excel:
        with open(arquivo_excel, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application', subtype='xlsx', filename=arquivo_excel)
    if arquivo_pdf:
        with open(arquivo_pdf, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application', subtype='pdf', filename=f"DO_MPRJ_{data_do.replace('/','-')}.pdf")

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_REMETENTE, SENHA_APP)
            smtp.send_message(msg)
        print("E-mail formatado enviado com sucesso.")
    except Exception as e:
        print(f"Erro no envio do e-mail: {e}")

def rodar():
    data_alvo = "08.05.2026"
    data_exibicao = "08/05/2026"
    url_pdf = f"https://www.mprj.mp.br/documents/20184/8887328/{data_alvo}.pdf"
    
    localizado = False
    status_download = "Nao iniciado"
    status_ia = "Nao iniciado"
    tem_dados = False
    tamanho_pdf_kb = 0
    
    try:
        print(f"Buscando PDF: {url_pdf}")
        response = requests.get(url_pdf, timeout=30)
        
        if response.status_code == 200:
            localizado = True
            status_download = "Bem sucedido"
            
            pdf_local = "temp_diario.pdf"
            with open(pdf_local, "wb") as f:
                f.write(response.content)

            tamanho_pdf_kb = round(os.path.getsize(pdf_local) / 1024, 2)

            dados, status_ia, tempo_processamento = extrair_dados_com_ia(pdf_local)

            if dados:
                tem_dados = True
                qtd_vagas = len(dados)
                excel_local = "Vagas_Encontradas.xlsx"
                formatar_excel(dados, excel_local, data_exibicao)
                
                enviar_email(data_exibicao, url_pdf, localizado, status_download, status_ia, tem_dados, 
                             qtd_vagas=qtd_vagas, tempo_ia=tempo_processamento, tamanho_kb=tamanho_pdf_kb, 
                             arquivo_excel=excel_local, arquivo_pdf=pdf_local)
            else:
                enviar_email(data_exibicao, url_pdf, localizado, status_download, status_ia, tem_dados, 
                             tempo_ia=tempo_processamento, tamanho_kb=tamanho_pdf_kb, arquivo_pdf=pdf_local)
        
        else:
            status_download = f"Mal sucedido (Erro {response.status_code})"
            enviar_email(data_exibicao, url_pdf, localizado, status_download, status_ia, tem_dados)
            
    except Exception as e:
        status_download = f"Mal sucedido ({str(e)})"
        print(f"Erro critico: {e}")
        enviar_email(data_exibicao, url_pdf, localizado, status_download, status_ia, tem_dados)

if __name__ == "__main__":
    rodar()
